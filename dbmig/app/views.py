
from django.shortcuts import render
import openpyxl
from django.core.files.storage import FileSystemStorage
from app.forms import UserdetailsCreationForm, FormUploadFileData, LrtransationForm
from django.views.generic.edit import FormView, View, CreateView
from tablib import Dataset
from app.models import Lrgeneratingmtbl, Companywarehousemaster, Cosingneemaster, Userdetails, Lrdocument, Lrtransation
from django.urls import reverse_lazy,reverse
from django.contrib import messages
from django.contrib.auth import (authenticate, login, logout,
                                 update_session_auth_hash)
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import (AuthenticationForm, UserCreationForm,
                                       PasswordChangeForm)
from django.http import HttpResponseRedirect
from django.template.loader import render_to_string
from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic.base import TemplateView
from . import models
from django.contrib import auth
from . import forms
from app.forms import UserdetailsCreationForm, CompanywarehouseForm
from django.views.generic import View
from io import BytesIO, StringIO
from django.db.models import Q
from reportlab.pdfgen import canvas
from django.http import HttpResponse
from django.conf import settings
from django.views.generic import DetailView, UpdateView
import django_weasyprint
from django_weasyprint import WeasyTemplateResponseMixin
from weasyprint import HTML
import tempfile
from xhtml2pdf import pisa 
from django.template.loader import get_template 
from django.template import Context 

def html_to_pdf_directly(request): 
    template = get_template("lrprint.html") 
    context = {'pagesize':'A4'} 
    html = template.render(context) 
    result = BytesIO() 
    pdf = pisa.pisaDocument(StringIO(html), dest=result) 
    if not pdf.err: 
        return HttpResponse(result.getvalue(), content_type='application/pdf') 
    else: return HttpResponse('Errors') 


class MyModelView(DetailView):
    # vanilla Django DetailView
    model = Lrgeneratingmtbl
    template_name = 'lrprint.html'

class MyModelPrintView(WeasyTemplateResponseMixin, MyModelView):
    # output of DetailView rendered as PDF
    pdf_stylesheets = [
        settings.STATIC_ROOT,
    ]
    # show pdf in-line (default: True, show download dialog)
    pdf_attachment = False
    # suggested filename (is required for attachment!)
    pdf_filename = 'lrcopies.pdf'

    # def get_object(self):
    #     return get_object_or_404(Userdetails, pk=self.request.user.id)

def generate_pdf(request):
    """Generate pdf."""
    # Model data
    lrdtls = Lrgeneratingmtbl.objects.all()

    # Rendered
    html_string = render_to_string('lrprint.html', {'lrdtls': lrdtls})
    html = HTML(string=html_string)
    result = html.write_pdf()

    # Creating http response
    response = HttpResponse(content_type='application/pdf;')
    response['Content-Disposition'] = 'inline; filename=lrcopies.pdf'
    response['Content-Transfer-Encoding'] = 'binary'
    with tempfile.NamedTemporaryFile(delete=True) as output:
        output.write(result)
        output.flush()
        output = open(output.name,'rb')
        response.write(output.read())

    return response

def deletelr(request):
   return render(request, "app/lrcancel.html", {})

def duplicate(request):
   return render(request, "app/lrduplct.html", {})

def change(request):
   return render(request, "app/lrchange.html", {})

def editvno(request):
   return render(request, "app/vehicleno.html", {})



class LogoutView(View):
    def get(self, request):
        logout(request)
        form = UserdetailsCreationForm(self.request.POST or None)
        context = {
            "form": form,
        }
        return render(self.request, 'sign_in.html', context)


# class LoginView(View):
#     def post(self, request):
#         username = request.POST['username']
#         password = request.POST['password']
#         user = authenticate(usrd_name=username, password=password)
#         if user is not None:
#             if user.is_active:
#                 login(request, user)
#                 # whms = Companywarehousemaster.objects.all()
#                 return render(self.request, 'dash.html', {'user': user})
#             else:
#                 return render(self.request, 'sign_in.html', {'error_message': 'Your account has been disabled'})
#         else:
#             return render(self.request, 'sign_in.html', {'error_message': 'Invalid login'})
#         return render(self.request, 'sign_in.html')


def sign_in(request):
    form = AuthenticationForm()
    if request.method == 'POST':
        form = AuthenticationForm(data=request.POST)
        if form.is_valid():
            if form.user_cache is not None:
                user = form.user_cache
                if user.is_active:
                    auth.login(request, user)
                    return HttpResponseRedirect(
                        reverse('app:dash')  # TODO: go to profile
                    )
                else:
                    messages.error(
                        request,
                        "That user account has been disabled."
                    )
            else:
                messages.error(
                    request,
                    "Username or password is incorrect."
                )   
    return render(request, 'sign_in.html', {'form': form})


def register(request):
    form = UserdetailsCreationForm()
    if request.method == 'POST':
        form = UserdetailsCreationForm(data=request.POST)
        if form.is_valid():
            form.save()
            user = authenticate(
                username=form.cleaned_data['usrd_name'],
                password=form.cleaned_data['password']
            )
            login(request, user)
            messages.success(
                request,
                "You're now a user! You've been signed in, too."
            )
            return HttpResponseRedirect(reverse('app:dash'))  # TODO: go to profile
    return render(request, 'register.html', {'form': form})

# def signout(request):
#     auth.logout(request)
#     messages.success(request, "You've been signed out.")
#     return HttpResponseRedirect(reverse('app:signin'))

def profile(request):
    """Display User Profile"""
    profile = request.user.profile
    return render(request, 'profile.html', {
        'profile': profile
    })

# @render_to('dash.html') 
def dash(request):
    if not request.user.is_authenticated:
        return render(request, 'sign_in.html')
    else:
        base_template = 'base_%s.html' % request.user.usrd_usrcatpntr.usrcat_code
        return render(request, base_template)
    # else:
    #     return render(request, 'dash.html')

# def whm(request):
#     if not request.user.is_authenticated:
#         user = request.user
#         return render(request, 'sign_in.html')
#     else:
#         form = CompanywarehouseForm
#         return render(request, 'whm.html', {'form':form})

class WhmCreate(CreateView):

    model = Companywarehousemaster 
    template_name = 'whm.html'
    success_url = reverse_lazy('app:dash')
    fields = ('com_wmasname','com_wmasdesc','com_wmasaddress','com_wmasactive','com_wmasremarks')

def pod(request):
    lrdtls_results = Lrtransation.objects.all()
    # song_results = Song.objects.all()
    query = request.GET.get("q")
    if query:
        lrdtls_results = lrdtls_results.filter(
            Q(lrtran_frtypebillno__icontains=query)
        ).distinct()
        return render(request, 'poddisp.html', {
            'lrdtls': lrdtls_results,
        })
    else:
        return render(request, 'poddisp.html', {'lrdtls_results': lrdtls_results})

class LrdocumentCreate(CreateView):

    model = Lrdocument 
    template_name = 'podc.html'
    success_url = reverse_lazy('app:dash')
    fields = ('lrdoc_data','lrdoc_remarks')

    # if request.method == 'POST' and request.FILES['myfile']:
    #     myfile = request.FILES['myfile']
    #     fs = FileSystemStorage()
    #     filename = fs.save(myfile.name, myfile)
    #     uploaded_file_url = fs.url(filename)
    #     return render(request, 'podc.html', {
    #         'uploaded_file_url': uploaded_file_url
    #     })
    # return render(request, 'podc.html')

def pod_data(request,**kwargs):

    id = kwargs['_id']
    tlink = Lrtranslink.objects.filter(id=id).first()
    songs = tlink.songs.all()
    return render(request, 'pod_data.html', {
        'song_list': songs,
    })

# def lr(request):
#     message=''
#     if request.method == 'POST':
#         form = FormUploadFileData(request.POST, request.FILES)
#         if form.is_valid():
#             from projects.models import Project
#             excel_file = request.FILES['excel_file']
#             try:
#                 import os
#                 import tempfile
#                 import xlrd
#                 fd, tmp = tempfile.mkstemp() # create two temporary file
#                 with os.open(fd, 'wb') as out: # create new file objects
#                     out.write(excel_file.read())
#                 book = xlrd.open_workbook(fd)
#                 sheet = book.sheet_by_index(0)
#                 obj=Project(
#                     lrg_tpcode = sheet.cell_value(rowx=1, colx=1),
#                     lrg_shiptocode = sheet.cell_value(rowx=3, colx=1),
#                     lrg_pl = sheet.cell_value(rowx=4, colx=1),
#                     lrg_div = sheet.cell_value(rowx=5, colx=1),
#                     lrg_matgrp = sheet.cell_value(rowx=6, colx=1),
#                     lrg_material = sheet.cell_value(rowx=8, colx=1),
#                     lrg_hsncode = sheet.cell_value(rowx=9, colx=1),
#                     lrg_taxinvno = sheet.cell_value(rowx=12, colx=1),
#                     lrg_intinvno = sheet.cell_value(rowx=13, colx=1),
#                     lrg_date = sheet.cell_value(rowx=14, colx=1),
#                     lrg_qty = sheet.cell_value(rowx=15, colx=1),
#                 )
                
#                 obj.save()
#             finally:
#                 os.unlink(tmp)
#         else:
#             message='Invalid Entries'
#     else:
#         form = FormUploadFileData()
#     return render(request,'lr.html', {'form':form,'message':message})

 
class LorryReceiptView(FormView):
    template_name = 'lr.html'
    form_class = LrtransationForm
    success_url = reverse_lazy('app:validate')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # context["vehicle_types"] =  
        # import pdb;pdb.set_trace()
        return context
    
    def form_valid(self, form):
        import xlrd
        import datetime
        excel_file = self.request.FILES["excel_file"]

        # you may put validations here to check extension or file size
        # book = xlrd.open_workbook(excel_file)
        # sheet = book.sheet_by_index(0)
    
        wb = openpyxl.load_workbook(excel_file)

        # getting a particular sheet by name out of many sheets
        sheet = wb["Sheet2"]
        print(sheet)

        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in sheet.iter_rows(min_row=2):
            row_data = excel_data
            if row[0].value is None:
                    break
            else:
                row_data = {
                'lrg_tpcode' :str(row[0].value),
                'lrg_shiptocode': str(row[2].value),
                'lrg_pl': str(row[3].value),
                'lrg_div': str(row[4].value),
                'lrg_matgrp' : str(row[5].value),
                "lrg_material": str(row[7].value),
                "lrg_hsncode" : str(row[8].value),
                "lrg_taxinvno" : str(row[11].value),
                "lrg_intinvno" : str(row[12].value),
                "lrg_date" : datetime.datetime.strptime(row[13].value, '%d.%m.%Y'),
                "lrg_qty" : row[14].value,
                }
            if row_data != {}:
                excel_data.append(row_data)
        from pprint import pprint
        pprint(excel_data)
        for row in excel_data:
            Lrgeneratingmtbl.objects.create(**row)
        # instance = form.save()
        return super().form_valid(form)



# def upload(request):
#     form = None
#     if request.method == "POST":
#         form = UploadFileForm(request.POST, request.FILES)
#         if form.is_valid():
#             filehandle = request.FILES['file']
#             # call import_sheet to import the sheet into your database
#             return import_sheet(request)  
#     form = form or UploadFileForm()
#     return render(request,'upload_form.html',{
#         'form': form,
#         'title': 'Excel file upload',
#         'header': 'Please choose a valid excel file'
#     })

#
# def lr(request):
#     if "GET" == request.method:
#         return render(request, 'lr.html', {})
#     else:
#         excel_file = request.FILES["excel_file"]

#         # you may put validations here to check extension or file size

#         wb = openpyxl.load_workbook(excel_file)

#         # getting a particular sheet by name out of many sheets
#         worksheet = wb["Sheet1"]
#         print(worksheet)

#         excel_data = list()
#         # iterating over the rows and
#         # getting value from each cell in row
#         for row in worksheet.iter_rows():
#             row_data = list()
#             for cell in row:
#                 row_data.append(str(cell.value))
#             excel_data.append(row_data)
        
#         return render(request, 'lr.html', {"excel_data":excel_data})

def validate(request):
    query_results = Lrgeneratingmtbl.objects.all()

    return render(request,'validate.html', {'query_results':query_results})



# def print(request):
#     query_results = Lrgeneratingmtbl.objects.filter()

#     return render(request,'lr.html', {'query_results':query_results})
    # if request.method == 'POST':
    #     lrgeneratingmtbl_resource = LrgeneratingmtblResource()
    #     dataset = Dataset()
    #     new_ = request.FILES['myfile']

    #     imported_data = dataset.load(new_persons.read())
    #     result = lrgeneratingmtbl_resource.import_data(dataset, dry_run=True)  

    #     if not result.has_errors():
    #         lrgeneratingmtbl_resource.import_data(dataset, dry_run=False)  

    # return render(request, 'lr.html')



